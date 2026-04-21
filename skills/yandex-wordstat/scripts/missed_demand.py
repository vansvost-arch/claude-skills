#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""Missed demand analysis for Yandex Direct campaigns.

Subcommands:
  parse-xlsx   Parse Yandex Direct XLSX export into groups/phrases/minus
  build-query  Build OR-query from slot structure
  merge-slots  Merge batch segmentation results into unified slots
  query-total  Get totalCount from Wordstat API
"""
import argparse
import json
import sys
import urllib.request
import urllib.error
import re
from collections import OrderedDict


# --- Stop words that need + prefix in Wordstat queries ---
STOP_WORDS = frozenset([
    "на", "в", "к", "за", "с", "по", "из", "от", "до", "для",
    "без", "при", "под", "над", "между", "через", "об", "перед",
])

# Characters forbidden in slot variants (OR-syntax operators and leading modifiers)
# Hyphen inside words is OK (б/у, санкт-петербург)
FORBIDDEN_CHARS_RE = re.compile(r'[|()"]')
LEADING_OPERATOR_RE = re.compile(r'^[+\-!]')
# Punctuation to strip from slot variants (dots, commas, semicolons, etc.)
# Keeps hyphens (б/у, санкт-петербург), slashes, and ! (Wordstat exact form operator)
PUNCTUATION_RE = re.compile(r'[.,;:?…·•]')


def cmd_parse_xlsx(args):
    """Parse Yandex Direct XLSX export."""
    import openpyxl

    wb = openpyxl.load_workbook(args.file, data_only=True)
    if "Тексты" not in wb.sheetnames:
        sys.exit(
            f"Error: лист 'Тексты' не найден. Доступные: {wb.sheetnames}"
        )
    ws = wb["Тексты"]

    # Find campaign minus-phrases (scan rows 1-20, cols A-F)
    campaign_minus = ""
    found_minus = False
    for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=6, values_only=False):
        if found_minus:
            break
        for cell in row:
            if cell.value and "Минус-фразы на кампанию" in str(cell.value):
                next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                if next_cell.value:
                    campaign_minus = str(next_cell.value).strip()
                found_minus = True
                break

    # Find first data row: col A == "-" and col G non-empty
    data_start = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1, values_only=False):
        cell = row[0]
        if cell.value and str(cell.value).strip() == "-":
            # Check col G
            g_cell = ws.cell(row=cell.row, column=7)
            if g_cell.value and str(g_cell.value).strip():
                data_start = cell.row
                break

    if data_start is None:
        sys.exit("Error: не найдены строки данных (col A='-' с непустой col G)")

    # Collect groups
    groups = OrderedDict()  # group_id -> {name, group_minus_set, phrases_list}
    for row_idx in range(data_start, ws.max_row + 1):
        col_a = ws.cell(row=row_idx, column=1).value
        if col_a is None or str(col_a).strip() != "-":
            continue

        col_g = ws.cell(row=row_idx, column=7).value
        if col_g is None or not str(col_g).strip():
            continue

        col_c = ws.cell(row=row_idx, column=3).value
        if col_c is None or not str(col_c).strip():
            continue  # skip rows with empty group_id

        group_id = str(col_c).strip()
        group_name = str(ws.cell(row=row_idx, column=4).value or "").strip()
        phrase = str(col_g).strip()

        # Col AG (33) - group minus
        col_ag = ws.cell(row=row_idx, column=33).value
        group_minus_val = str(col_ag).strip() if col_ag else ""

        if group_id not in groups:
            groups[group_id] = {
                "name": group_name,
                "group_minus_set": set(),
                "phrases_ordered": OrderedDict(),
            }

        g = groups[group_id]
        # Deduplicate phrases preserving order
        g["phrases_ordered"][phrase] = None
        if group_minus_val:
            g["group_minus_set"].add(group_minus_val)

    wb.close()

    # Build output
    result_groups = []
    for gid, g in groups.items():
        result_groups.append({
            "id": gid,
            "name": g["name"],
            "group_minus": " ".join(sorted(g["group_minus_set"])),
            "phrases": list(g["phrases_ordered"].keys()),
        })

    # Filter by --group if specified
    if args.group:
        result_groups = [g for g in result_groups if g["id"] == args.group]

    result = {
        "campaign_minus": campaign_minus,
        "campaign_minus_length": len(campaign_minus),
        "groups": result_groups,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


def sanitize_variant(variant, slot_name):
    """Sanitize a slot variant. Returns (cleaned, warnings)."""
    warnings = []
    cleaned = variant

    # Strip punctuation (муз. центр -> муз центр)
    if PUNCTUATION_RE.search(cleaned):
        cleaned = PUNCTUATION_RE.sub("", cleaned)
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()

    # Remove forbidden OR-syntax characters (replace with space to avoid merging tokens)
    if FORBIDDEN_CHARS_RE.search(cleaned):
        warnings.append(
            f"slot '{slot_name}': вариант '{variant}' содержит OR-синтаксис, деградирован в многословный"
        )
        cleaned = FORBIDDEN_CHARS_RE.sub(" ", cleaned)
        # Collapse multiple spaces
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()

    # Remove leading operators (+ - !) from each token
    tokens = cleaned.split()
    new_tokens = []
    for t in tokens:
        if LEADING_OPERATOR_RE.match(t):
            stripped = LEADING_OPERATOR_RE.sub("", t)
            if stripped:
                warnings.append(
                    f"slot '{slot_name}': убран ведущий оператор из '{t}'"
                )
                new_tokens.append(stripped)
        else:
            new_tokens.append(t)
    cleaned = " ".join(new_tokens)

    return cleaned.strip(), warnings


def add_stop_word_plus(variant):
    """Add + prefix to stop words in a variant string."""
    tokens = variant.split()
    result = []
    for t in tokens:
        # Don't touch tokens already starting with operator
        if t[0] in ("+", "-", "!"):
            result.append(t)
        elif t.lower() in STOP_WORDS:
            result.append(f"+{t}")
        else:
            result.append(t)
    return " ".join(result)


def normalize_token(token):
    """Normalize a token for comparison: lowercase, strip punctuation."""
    t = token.lower().strip()
    t = PUNCTUATION_RE.sub("", t)
    t = LEADING_OPERATOR_RE.sub("", t)
    t = FORBIDDEN_CHARS_RE.sub("", t)
    return t.strip()


class SlotMerger:
    """Merges slot segmentation results from multiple LLM batches."""

    SLOT_NAMES = ["objects", "actions", "modifiers", "additional"]
    ADDITIONAL_PATTERN_RE = re.compile(
        r'\b(' + '|'.join(STOP_WORDS) + r')\s+(\S+)',
        re.IGNORECASE,
    )

    def __init__(self, phrases):
        self._phrases = list(phrases)
        self._slots = {
            name: OrderedDict() for name in self.SLOT_NAMES
        }

    def merge_batch(self, batch_index, phrase_indexes, slots):
        """Merge slots from one batch into accumulated state."""
        n = len(self._phrases)
        for pi in phrase_indexes:
            if not isinstance(pi, int) or pi < 0 or pi >= n:
                raise ValueError(
                    f"phrase_index {pi} вне диапазона [0..{n - 1}]"
                )
        for slot_name in self.SLOT_NAMES:
            variants = slots.get(slot_name, [])
            if not isinstance(variants, list):
                continue
            for v in variants:
                v_str = str(v).strip()
                if not v_str:
                    continue
                cleaned, _ = sanitize_variant(v_str, slot_name)
                if not cleaned:
                    continue
                canon = re.sub(r'\s+', ' ', cleaned.lower()).strip()
                slot_dict = self._slots[slot_name]
                if canon not in slot_dict:
                    slot_dict[canon] = {
                        "display": cleaned,
                        "batch_indexes": set(),
                    }
                slot_dict[canon]["batch_indexes"].add(batch_index)

    def _remove_subsets(self):
        """Remove subset variants within each slot (heuristic)."""
        for slot_name in self.SLOT_NAMES:
            slot_dict = self._slots[slot_name]
            keys = list(slot_dict.keys())
            token_sets = {}
            for k in keys:
                tokens = set(
                    normalize_token(t)
                    for t in k.split()
                    if normalize_token(t)
                )
                token_sets[k] = tokens
            to_remove = set()
            for i, k_short in enumerate(keys):
                for j, k_long in enumerate(keys):
                    if i == j or k_long in to_remove:
                        continue
                    ts_short = token_sets[k_short]
                    ts_long = token_sets[k_long]
                    if ts_short and ts_short < ts_long:
                        to_remove.add(k_long)
            for k in to_remove:
                del slot_dict[k]

    def _compute_coverage(self):
        """Compute coverage of original phrases by merged slot variants."""
        all_variant_tokens = set()
        for slot_name in self.SLOT_NAMES:
            for canon in self._slots[slot_name]:
                for t in canon.split():
                    nt = normalize_token(t)
                    if nt:
                        all_variant_tokens.add(nt)

        uncovered_phrases = []
        all_phrase_tokens = set()
        for idx, phrase in enumerate(self._phrases):
            phrase_tokens = set()
            for t in phrase.split():
                nt = normalize_token(t)
                if nt and nt not in STOP_WORDS:
                    phrase_tokens.add(nt)
                    all_phrase_tokens.add(nt)
            if phrase_tokens and not (phrase_tokens & all_variant_tokens):
                uncovered_phrases.append({"index": idx, "phrase": phrase})

        uncovered_tokens = sorted(
            all_phrase_tokens - all_variant_tokens - STOP_WORDS
        )

        additional_variants_lower = set(
            self._slots.get("additional", OrderedDict()).keys()
        )
        additional_patterns = []
        seen_patterns = set()
        for phrase in self._phrases:
            normalized_phrase = PUNCTUATION_RE.sub("", phrase.lower())
            normalized_phrase = re.sub(r'\s+', ' ', normalized_phrase).strip()
            for m in self.ADDITIONAL_PATTERN_RE.finditer(normalized_phrase):
                pattern = re.sub(r'\s+', ' ', m.group(0)).strip()
                if pattern not in seen_patterns:
                    seen_patterns.add(pattern)
                    if pattern not in additional_variants_lower:
                        additional_patterns.append(pattern)

        return {
            "uncovered_phrases": uncovered_phrases,
            "uncovered_tokens": uncovered_tokens,
            "additional_patterns": additional_patterns,
        }

    def finalize(self, max_query_length=4096):
        """Finalize merged slots: deduplicate, remove subsets, build query."""
        self._remove_subsets()

        slots_pre_trim = {}
        for slot_name in self.SLOT_NAMES:
            slots_pre_trim[slot_name] = [
                entry["display"]
                for entry in self._slots[slot_name].values()
            ]

        coverage = self._compute_coverage()

        debug = {}
        for slot_name in self.SLOT_NAMES:
            debug[slot_name] = {
                entry["display"]: sorted(entry["batch_indexes"])
                for entry in self._slots[slot_name].values()
            }

        slot_order = self.SLOT_NAMES
        warnings = []
        trimmed = []

        sanitized_slots = {}
        for slot_name in slot_order:
            clean_variants = []
            for entry in self._slots[slot_name].values():
                clean_variants.append(add_stop_word_plus(entry["display"]))
            sanitized_slots[slot_name] = clean_variants

        def build_or_string(slot_variants):
            if not slot_variants:
                return ""
            if len(slot_variants) == 1:
                return slot_variants[0]
            return "(" + "|".join(slot_variants) + ")"

        def assemble_query(s_slots):
            parts = []
            for name in slot_order:
                or_str = build_or_string(s_slots[name])
                if or_str:
                    parts.append(or_str)
            return " ".join(parts)

        def est_phrases(s_slots):
            product = 1
            for name in slot_order:
                n = len(s_slots[name])
                if n > 0:
                    product *= n
            return product

        trim_order = ["additional", "modifiers", "actions"]
        while est_phrases(sanitized_slots) > 200:
            trimmed_any = False
            for trim_slot in trim_order:
                if len(sanitized_slots[trim_slot]) > 1:
                    removed = sanitized_slots[trim_slot].pop()
                    trimmed.append({
                        "slot": trim_slot, "removed": removed,
                        "reason": "estimated_phrases > 200",
                    })
                    trimmed_any = True
                    break
            if not trimmed_any:
                break

        query = assemble_query(sanitized_slots)
        while len(query) > max_query_length:
            trimmed_any = False
            for trim_slot in trim_order:
                if len(sanitized_slots[trim_slot]) > 1:
                    removed = sanitized_slots[trim_slot].pop()
                    trimmed.append({
                        "slot": trim_slot, "removed": removed,
                        "reason": f"query_length > {max_query_length}",
                    })
                    trimmed_any = True
                    query = assemble_query(sanitized_slots)
                    break
            if not trimmed_any:
                warnings.append(
                    f"query_length {len(query)} > {max_query_length}"
                    " после максимальной обрезки"
                )
                break

        slots_post_trim = {}
        for slot_name in slot_order:
            slots_post_trim[slot_name] = list(sanitized_slots[slot_name])

        return {
            "slots_pre_trim": slots_pre_trim,
            "slots_post_trim": slots_post_trim,
            "query": query,
            "estimated_phrases": est_phrases(sanitized_slots),
            "query_length": len(query),
            "trimmed": trimmed,
            "warnings": warnings,
            "coverage": coverage,
            "debug": debug,
        }


def cmd_merge_slots(args):
    """Merge batch segmentation results from stdin JSON."""
    raw = sys.stdin.read()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        sys.exit(f"Error: невалидный JSON на stdin: {e}")

    phrases = data.get("phrases", [])
    batches = data.get("batches", [])

    if not isinstance(phrases, list) or not phrases:
        sys.exit("Error: phrases должен быть непустым списком строк")
    if not isinstance(batches, list) or not batches:
        sys.exit("Error: batches должен быть непустым списком объектов")
    for p in phrases:
        if not isinstance(p, str):
            sys.exit(f"Error: элемент phrases не строка: {type(p).__name__}")

    merger = SlotMerger(phrases)
    for i, batch in enumerate(batches):
        if not isinstance(batch, dict):
            sys.exit(f"Error: batch[{i}] не объект")
        phrase_indexes = batch.get("phrase_indexes", [])
        if not isinstance(phrase_indexes, list):
            sys.exit(f"Error: batch[{i}].phrase_indexes не список")
        slots = batch.get("slots", {})
        if not isinstance(slots, dict):
            sys.exit(f"Error: batch[{i}].slots не объект")
        try:
            merger.merge_batch(i, phrase_indexes, slots)
        except ValueError as e:
            sys.exit(f"Error: batch[{i}]: {e}")

    result = merger.finalize(max_query_length=args.max_query_length)
    print(json.dumps(result, ensure_ascii=False, indent=2))


def cmd_build_query(args):
    """Build OR-query from slots JSON."""
    try:
        slots = json.loads(args.slots_json)
    except json.JSONDecodeError as e:
        sys.exit(f"Error: невалидный JSON слотов: {e}")

    max_len = args.max_query_length
    warnings = []
    trimmed = []

    slot_order = ["actions", "objects", "modifiers", "additional"]

    sanitized_slots = {}
    for slot_name in slot_order:
        raw_variants = slots.get(slot_name, [])
        if not isinstance(raw_variants, list):
            raw_variants = []
        clean_variants = []
        for v in raw_variants:
            v_str = str(v).strip()
            if not v_str:
                continue
            cleaned, san_warnings = sanitize_variant(v_str, slot_name)
            warnings.extend(san_warnings)
            if cleaned:
                cleaned = add_stop_word_plus(cleaned)
                clean_variants.append(cleaned)
        sanitized_slots[slot_name] = clean_variants

    def build_or_string(slot_variants):
        if len(slot_variants) == 0:
            return ""
        if len(slot_variants) == 1:
            return slot_variants[0]
        return "(" + "|".join(slot_variants) + ")"

    def assemble_query(s_slots):
        parts = []
        for name in slot_order:
            or_str = build_or_string(s_slots[name])
            if or_str:
                parts.append(or_str)
        return " ".join(parts)

    def estimated_phrases(s_slots):
        product = 1
        for name in slot_order:
            n = len(s_slots[name])
            if n > 0:
                product *= n
        return product

    trim_order = ["additional", "modifiers", "actions"]

    while estimated_phrases(sanitized_slots) > 200:
        trimmed_any = False
        for trim_slot in trim_order:
            if len(sanitized_slots[trim_slot]) > 1:
                removed = sanitized_slots[trim_slot].pop()
                trimmed.append({"slot": trim_slot, "removed": removed, "reason": "estimated_phrases > 200"})
                trimmed_any = True
                break
        if not trimmed_any:
            break

    query = assemble_query(sanitized_slots)
    while len(query) > max_len:
        trimmed_any = False
        for trim_slot in trim_order:
            if len(sanitized_slots[trim_slot]) > 1:
                removed = sanitized_slots[trim_slot].pop()
                trimmed.append({"slot": trim_slot, "removed": removed, "reason": f"query_length > {max_len}"})
                trimmed_any = True
                query = assemble_query(sanitized_slots)
                break
        if not trimmed_any:
            warnings.append(
                f"query_length {len(query)} всё ещё > {max_len} после максимальной обрезки"
            )
            break

    result = {
        "query": query,
        "estimated_phrases": estimated_phrases(sanitized_slots),
        "query_length": len(query),
        "trimmed": trimmed,
        "warnings": warnings,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


def cmd_query_total(args):
    """Get totalCount from Wordstat API."""
    body = {"phrase": args.phrase}
    if args.regions:
        try:
            body["regions"] = [int(r.strip()) for r in args.regions.split(",")]
        except ValueError:
            sys.exit(f"Error: невалидные regions: {args.regions}")

    data = json.dumps(body).encode("utf-8")
    url = "https://api.wordstat.yandex.net/v1/topRequests"

    req = urllib.request.Request(
        url,
        data=data,
        headers={
            "Authorization": f"Bearer {args.token}",
            "Content-Type": "application/json; charset=utf-8",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            resp_body = resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        raw = ""
        try:
            raw = e.read().decode("utf-8")[:500]
        except Exception:
            pass
        result = {"error": f"HTTP {e.code}: {e.reason}", "query": args.phrase, "raw": raw}
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)
    except urllib.error.URLError as e:
        result = {"error": f"URL error: {e.reason}", "query": args.phrase}
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)

    try:
        obj = json.loads(resp_body)
    except json.JSONDecodeError:
        result = {"error": "Invalid JSON response", "query": args.phrase, "raw": resp_body[:500]}
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)

    if "error" in obj:
        err_val = obj["error"]
        result = {"error": str(err_val), "query": args.phrase, "raw": resp_body[:500]}
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)

    if "error_code" in obj:
        err_msg = f"{obj.get('error_str', '')}: {obj.get('error_detail', '')}"
        result = {
            "error": err_msg.strip(": "),
            "error_code": obj["error_code"],
            "query": args.phrase,
            "raw": resp_body[:500],
        }
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)

    total_count = obj.get("totalCount")
    if total_count is None:
        total_count = (obj.get("result") or {}).get("totalCount")
    if total_count is None:
        result = {
            "error": "totalCount not found in response",
            "query": args.phrase,
            "raw": resp_body[:500],
        }
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)

    result = {"total_count": int(total_count), "query": args.phrase}
    print(json.dumps(result, ensure_ascii=False))


def main():
    parser = argparse.ArgumentParser(description="Missed demand analysis tools")
    sub = parser.add_subparsers(dest="command", required=True)

    # parse-xlsx
    p_parse = sub.add_parser("parse-xlsx", help="Parse Yandex Direct XLSX export")
    p_parse.add_argument("file", help="Path to XLSX file")
    p_parse.add_argument("--group", default=None, help="Filter by group ID")

    # build-query
    p_build = sub.add_parser("build-query", help="Build OR-query from slots")
    p_build.add_argument("slots_json", help="JSON string with slots")
    p_build.add_argument("--max-query-length", type=int, default=4096, help="Max query length (default: 4096)")

    # merge-slots
    p_merge = sub.add_parser(
        "merge-slots",
        help="Merge batch segmentation results (stdin JSON)",
    )
    p_merge.add_argument(
        "--max-query-length", type=int, default=4096,
        help="Max query length (default: 4096)",
    )

    # query-total
    p_query = sub.add_parser("query-total", help="Get totalCount from Wordstat API")
    p_query.add_argument("--token", required=True, help="Yandex Wordstat API token")
    p_query.add_argument("--phrase", required=True, help="Search phrase with operators")
    p_query.add_argument("--regions", default=None, help="Region IDs comma-separated")

    args = parser.parse_args()
    if args.command == "parse-xlsx":
        cmd_parse_xlsx(args)
    elif args.command == "build-query":
        cmd_build_query(args)
    elif args.command == "merge-slots":
        cmd_merge_slots(args)
    elif args.command == "query-total":
        cmd_query_total(args)


if __name__ == "__main__":
    main()
